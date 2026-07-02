"""Model Integrity Agent.

Audits an uploaded Excel financial model (openpyxl) for errors, hard-coded
values, and inconsistencies, then produces a narrative audit report via
Azure OpenAI.
"""

from config import require_env
from llm import llm


def model_integrity_agent_app():
    """
    A secure agent to analyze and audit Excel financial models for errors and inconsistencies.
    """
    # --- Local imports ---
    import io, re, html, markdown
    import streamlit as st
    import pandas as pd
    from openai import AzureOpenAI
    import openpyxl
    from openpyxl.utils import get_column_letter

    st.markdown("### 🛡️ Model Integrity Agent")
    st.markdown(
        "Audit confidential financial models with enterprise-grade privacy. Upload an Excel model to check for common errors, hard-coded values, and inconsistencies."
    )

    # --- AGENT CONFIG (Fetched from secrets for Azure) ---
    _cfg = require_env("AZURE_OPENAI_ENDPOINT", "AZURE_OPENAI_KEY", "AZURE_OPENAI_DEPLOYMENT_NAME")
    openai_endpoint = _cfg["AZURE_OPENAI_ENDPOINT"]
    openai_key = _cfg["AZURE_OPENAI_KEY"]
    openai_deployment_name = _cfg["AZURE_OPENAI_DEPLOYMENT_NAME"]

    # --- LOCAL HELPER FUNCTIONS ---
    def generate_report_html_from_markdown(analysis_results: dict) -> str:
        """
        Converts a dictionary of markdown analysis into a complete, styled HTML string.
        This helper is self-contained within the Model Integrity Agent.
        """
        report_title = "Financial Model Integrity Report"
        styles = """
        <style>
            .analysis-container { font-family: 'Poppins', sans-serif; border: 1px solid #e0e0e0; border-radius: 8px; padding: 25px; background-color: #f9fafb; margin: 20px; }
            .analysis-container h1 { font-size: 1.8em; font-weight: 700; color: #00416A; margin-top: 0; padding-bottom: 15px; border-bottom: 3px solid #00416A; }
            .analysis-container h2 { font-size: 1.5em; font-weight: 600; color: #00416A; border-bottom: 2px solid #e6f1f6; padding-bottom: 10px; margin-top: 30px; margin-bottom: 20px; }
            .analysis-container h3 { font-size: 1.2em; font-weight: 600; color: #1e1e1e; margin-top: 25px; margin-bottom: 10px; }
            .analysis-container p { margin-bottom: 1em; line-height: 1.6; color: #333; }
            .analysis-container ul, .analysis-container ol { list-style-position: outside; padding-left: 20px; margin-top: 1em; margin-bottom: 1em; }
            .analysis-container li { margin-bottom: 0.75em; line-height: 1.6; }
            .analysis-container table { width: 100%; border-collapse: collapse; margin: 20px 0; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }
            .analysis-container th, .analysis-container td { border: 1px solid #ddd; padding: 12px 15px; text-align: left; }
            .analysis-container th { background-color: #e6f1f6; font-weight: 600; }
            .analysis-container tr:nth-of-type(even) { background-color: #fdfdfd; }
        </style>
        """
        
        full_html_body = f"<h1>{html.escape(report_title)}</h1>"
        for title, markdown_content in analysis_results.items():
            full_html_body += f"<h2>{html.escape(title)}</h2>"
            html_from_md = markdown.markdown(markdown_content, extensions=['tables'])
            processed_html = re.sub(r"<h2>(.*?)</h2>", r"<h3>\1</h3>", html_from_md)
            full_html_body += processed_html
        
        content_div = f"<div class='analysis-container'>{full_html_body}</div>"
        
        return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><title>{html.escape(report_title)}</title>
        <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@400;600&display=swap" rel="stylesheet">
        {styles}</head><body>{content_div}</body></html>"""

    def audit_excel_model(file_bytes: bytes) -> dict:
        """
        Parses an Excel workbook and checks for common modeling errors.
        Returns a dictionary of findings.
        """
        findings = {
            "hard_codes": [],
            "error_cells": [],
            "external_links": [],
            "summary": {}
        }
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
            findings["summary"]["sheets"] = workbook.sheetnames
            findings["summary"]["total_sheets"] = len(workbook.sheetnames)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
                    for cell in row:
                        if cell.value is not None and cell.data_type == 'n' and not str(cell.value).startswith('='):
                            left_cell = sheet.cell(row=cell.row, column=cell.column - 1) if cell.column > 1 else None
                            if left_cell and left_cell.data_type == 'f':
                                findings["hard_codes"].append(
                                    f"Sheet '{sheet_name}', Cell {cell.coordinate}: Contains a hard-coded number ({cell.value}) next to a cell with a formula."
                                )
                        if cell.data_type == 'e':
                             findings["error_cells"].append(
                                f"Sheet '{sheet_name}', Cell {cell.coordinate}: Contains an error value: {cell.value}"
                            )
                if hasattr(sheet, 'external_links'):
                    for link in sheet.external_links:
                        findings["external_links"].append(f"Sheet '{sheet_name}' contains an external link to: {link.Target}")

            return findings
        except Exception as e:
            st.warning(f"Could not fully audit the Excel file. Error: {e}")
            return findings

    def analyze_findings_with_azure_openai(findings: dict) -> str:
        """
        Sends the structured findings to Azure OpenAI for a narrative report.
        """
        prompt = f"""
        You are an expert in financial modeling and auditing from a top-tier investment bank.
        An automated script has analyzed an Excel financial model and found the following potential issues.
        Your task is to synthesize these raw findings into a professional, well-structured audit report in MARKDOWN format.

        **CRITICAL INSTRUCTIONS:**
        1.  Start with a high-level executive summary based on the number and type of findings.
        2.  Group the findings into logical sections: "High-Severity Issues (e.g., Error Cells)", "Medium-Severity Issues (e.g., Hard-Codes in Calculation Blocks)", and "Areas for Review (e.g., External Links)".
        3.  For each finding, explain the potential risk (e.g., "Hard-coded values can lead to incorrect calculations if assumptions change and are a common source of model errors.").
        4.  Conclude with a summary and a clear recommendation for a manual review of the identified areas.
        5.  If a category (like 'error_cells') is empty, state that "No issues of this type were detected."

        **AUTOMATED FINDINGS:**
        ---
        {str(findings)}
        ---
        """
        try:
            return llm.chat(
                [
                    {"role": "system", "content": "You are a financial modeling audit expert."},
                    {"role": "user", "content": prompt},
                ],
                provider="azure",
                model=openai_deployment_name,
            )
        except Exception as e:
            return f"## Error\n\n**Error during Azure OpenAI analysis:** {e}"

    # --- UI & WORKFLOW ---
    st.subheader("1. Upload Confidential Financial Model")
    uploaded_file = st.file_uploader(
        "Upload an Excel Model (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        key="model_integrity_uploader",
    )

    if uploaded_file:
        if st.button("Audit Model", type="primary", use_container_width=True):
            with st.spinner("Auditing model structure and generating report..."):
                file_bytes = uploaded_file.getvalue()
                findings = audit_excel_model(file_bytes)
                analysis_report = analyze_findings_with_azure_openai(findings)
                st.session_state.model_integrity_results = {
                    "Model Audit Report": analysis_report
                }

    if "model_integrity_results" in st.session_state:
        st.success("✅ Model audit complete!")
        st.markdown("---")
        st.subheader("2. Download Report")
        
        full_html_for_download = generate_report_html_from_markdown(
            st.session_state.model_integrity_results
        )
        
        st.download_button(
            label="📥 Download Audit Report as HTML",
            data=full_html_for_download,
            file_name="model_integrity_report.html",
            mime="text/html",
            use_container_width=True
        )
